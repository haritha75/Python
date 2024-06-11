import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
wb.create_sheet("Sheet2",0)
# wb.remove_sheet(sheet)
cell = sheet["a1"]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# print(sheet.max_row)
# print(sheet.max_column)
# cell =  sheet.cell(row=1,column=1)

# iterate rows and columns

# for row in range(1,sheet.max_row+1):
#   for col in range(1,sheet.max_column+1):
#     cell = sheet.cell(row,col)
#     print(cell.value)

# column = sheet["a"]
# cells = sheet["a:c"] # it will take a to c cells
# print(cells)
# print(column)    

# print(sheet["a1:c3"])
# print(sheet[1]) #first row

sheet.append([1,2,3]) # to add row to the end of the sheet
# sheet.insert_rows()
# sheet.delete_rows()
wb.save("transaction2.xlsx")

